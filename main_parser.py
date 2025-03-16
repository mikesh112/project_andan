import schedule
import time
from datetime import datetime
import random
from openpyxl import load_workbook
import requests
from proxy_info import proxies
import pandas as pd

file_path = 'Цены ВБ.xlsx' # изначальный пустой файл, где были написаны артикулы, названия категорий и тд. Подготовительный файл для парсинга
main_df = pd.read_excel(file_path, skiprows=2)   



def get_wildberries_price(product_id):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://www.wildberries.ru/",
        "Connection": "keep-alive"
    } # пишем основную инфу про браузер и сайт, чтоб нормально спарсить

    api_url = f"https://card.wb.ru/cards/detail?appType=1&curr=rub&dest=-1257786&spp=0&nm={product_id}" # общая ссылка на продукт по артикулу

    try:
        response = requests.get(api_url, headers=headers, proxies=random.choice(proxies), timeout=10)
        response.raise_for_status()
        data = response.json() # респонсоб берем джейсон с сайта

        if "data" in data and "products" in data["data"] and data["data"]["products"]: # находим место, где находится цена товара
            price = data["data"]["products"][0].get("salePriceU", 0) // 100  # находим цену товара
            return price # возвращаем из ффункции цену, если нашли
        else:
            return "Цена не найдена" # обрабатываем если цена не найдена
    except requests.RequestException as e: 
        print("Ошибка при получении данных:", e) # обрабатываем ошибку если не получилось спарсить что-то
        return None



def main():
    current_time = datetime.now().replace(minute=0, second=0, microsecond=0)   # берем настоящее вреия системы, чтобы знать в какой момент была такая цена на товар
    print(f"Функция запущена в: {current_time}")
    prices = [] # массив цен
    
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S") # форматируем время, чтобы было удобно его сравнивать
    
    if datetime.strptime(main_df.iloc[-1, 0], '%Y-%m-%d %H:%M:%S') != \
            datetime.strptime(formatted_time, '%Y-%m-%d %H:%M:%S'): # проверяем, что еще не парсили эти данные
        
        for product_id in main_df.columns[1:]:  # собираем артикулы из изначального подготовительного файла
            price = get_wildberries_price(product_id) 
            prices.append(price) # добавляем цены в массив

        new_row = [formatted_time] + prices
        print(f"Новая строка добавлена: {new_row}") # вспомогательная шутка, чтобы в процессе парсинга смотреть, что все окей
        
        wb = load_workbook(file_path) # открываем исходный файлик
        ws = wb.active # нужно чтобы нормально вставлять и созранять  значения в ячейку
        next_row = ws.max_row + 1  # добавляем строчку (макс кол-во строчек)

        for col, value in enumerate(new_row, 1):
            ws.cell(row=next_row, column=col, value=value # добавляем в клетки найденые значения
        wb.save(file_path) # сохраняем файлик
                    
    else:
        print(f"Строка с датой {formatted_time} уже существует.")



schedule.every().hour.at(":01").do(main)  # каждый день, каждый час - расписание парсинга аххаха
while True:
    schedule.run_pending()
    time.sleep(600)  # задержка каждые 10 минут
